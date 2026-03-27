from io import BytesIO
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook

from carting.models import Cart, CartItem

from .models import Product, Vendor, VendorProductPrice
from .product_matching import normalize_product_name, product_matches_name


def _build_test_workbook(rows):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    for row in rows:
        worksheet.append(list(row))

    buffer = BytesIO()
    workbook.save(buffer)
    workbook.close()
    return buffer.getvalue()


class ProductAliasTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="inventory-admin",
            password="testpass123",
        )

    def setUp(self):
        self.client.force_login(self.user)

    def _product_payload(self, **overrides):
        payload = {
            "product_name": "Shrimp",
            "pack_type": "bag",
            "quantity_per_pack": "24",
            "quantity_unit": "G",
            "display_name": "Shrimp 24 G",
            "is_active": "on",
        }
        payload.update(overrides)
        return payload

    def test_add_product_saves_trimmed_non_empty_alternative_names(self):
        response = self.client.post(
            reverse("inventory:add_product"),
            data={
                **self._product_payload(),
                "alternative_names[]": [" Shrimp24G ", "", "Shrimp 24 G"],
            },
        )

        self.assertRedirects(response, reverse("inventory:manage_products"))

        product = Product.objects.get(display_name="Shrimp 24 G")
        self.assertEqual(
            list(product.aliases.order_by("alias_name").values_list("alias_name", flat=True)),
            ["Shrimp 24 G", "Shrimp24G"],
        )

    def test_edit_product_syncs_added_updated_and_removed_alternative_names(self):
        product = Product.objects.create(
            product_name="Shrimp",
            pack_type="bag",
            quantity_per_pack="24",
            quantity_unit="G",
            display_name="Shrimp 24 G",
            is_active=True,
        )
        product.aliases.create(alias_name="Shrimp24G")
        product.aliases.create(alias_name="Shrimp 24 Gram")

        response = self.client.post(
            reverse("inventory:edit_product", args=[product.pk]),
            data={
                **self._product_payload(display_name="Shrimp 24 G"),
                "alternative_names[]": ["Shrimp 24 G", " Shrimp Jumbo 24 G ", ""],
            },
        )

        self.assertRedirects(response, reverse("inventory:manage_products"))
        product.refresh_from_db()

        self.assertEqual(
            list(product.aliases.order_by("alias_name").values_list("alias_name", flat=True)),
            ["Shrimp 24 G", "Shrimp Jumbo 24 G"],
        )

    def test_product_matching_helper_checks_product_fields_and_aliases(self):
        product = Product.objects.create(
            product_name="Shrimp",
            pack_type="bag",
            quantity_per_pack="24",
            quantity_unit="G",
            display_name="Shrimp 24 G",
            is_active=True,
        )
        product.aliases.create(alias_name="Shrimp24G")

        self.assertEqual(normalize_product_name("  Shrimp   24 G  "), "shrimp 24 g")
        self.assertTrue(product_matches_name(product, " shrimp "))
        self.assertTrue(product_matches_name(product, "SHRIMP24G"))
        self.assertFalse(product_matches_name(product, "Prawn 24 G"))


class VendorPriceCurrencyTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="vendor-price-admin",
            password="testpass123",
        )
        cls.vendor = Vendor.objects.create(
            name="Euro Foods",
            email="euro@example.com",
            whatsapp_number="1234567890",
            is_active=True,
        )
        cls.product = Product.objects.create(
            product_name="Olive Oil",
            pack_type="tin",
            quantity_per_pack="5",
            quantity_unit="L",
            display_name="Olive Oil 5 L",
            is_active=True,
        )

    def setUp(self):
        self.client.force_login(self.user)

    def _vendor_price_payload(self, **overrides):
        payload = {
            "product": str(self.product.pk),
            "vendor": str(self.vendor.pk),
            "price": "50,25",
            "currency": "EUR",
            "is_active": "on",
        }
        payload.update(overrides)
        return payload

    def test_currency_dropdown_renders_supported_choices(self):
        response = self.client.get(reverse("inventory:add_vendor_price"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<option value="EUR">EUR — Euro</option>', html=True)
        self.assertContains(response, '<option value="USD">USD — US Dollar</option>', html=True)
        self.assertContains(response, '<option value="INR">INR — Indian Rupee</option>', html=True)
        self.assertContains(response, '<option value="CHF">CHF — Swiss Franc</option>', html=True)

    def test_eur_style_input_saves_decimal_price(self):
        response = self.client.post(
            reverse("inventory:add_vendor_price"),
            data=self._vendor_price_payload(price="50,25", currency="EUR"),
        )

        self.assertRedirects(response, reverse("inventory:manage_vendor_prices"))

        vendor_price = VendorProductPrice.objects.get()
        self.assertEqual(vendor_price.price, Decimal("50.25"))
        self.assertEqual(vendor_price.currency, "EUR")

    def test_usd_style_input_saves_decimal_price(self):
        response = self.client.post(
            reverse("inventory:add_vendor_price"),
            data=self._vendor_price_payload(price="50.25", currency="USD"),
        )

        self.assertRedirects(response, reverse("inventory:manage_vendor_prices"))

        vendor_price = VendorProductPrice.objects.get()
        self.assertEqual(vendor_price.price, Decimal("50.25"))
        self.assertEqual(vendor_price.currency, "USD")

    def test_invalid_mixed_input_is_rejected_safely(self):
        response = self.client.post(
            reverse("inventory:add_vendor_price"),
            data=self._vendor_price_payload(price="50,25.10", currency="EUR"),
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(
            response,
            "Enter a valid price using comma decimals, for example 50,25.",
        )
        self.assertFalse(VendorProductPrice.objects.exists())

    def test_form_and_vendor_price_list_display_reflect_selected_currency_format(self):
        vendor_price = VendorProductPrice.objects.create(
            product=self.product,
            vendor=self.vendor,
            price=Decimal("50.25"),
            currency="EUR",
            is_active=True,
        )
        usd_vendor = Vendor.objects.create(
            name="US Foods",
            email="us@example.com",
            whatsapp_number="9999999999",
            is_active=True,
        )
        usd_product = Product.objects.create(
            product_name="Tomato Sauce",
            pack_type="tin",
            quantity_per_pack="1",
            quantity_unit="L",
            display_name="Tomato Sauce 1 L",
            is_active=True,
        )
        VendorProductPrice.objects.create(
            product=usd_product,
            vendor=usd_vendor,
            price=Decimal("50.25"),
            currency="USD",
            is_active=True,
        )

        response = self.client.get(reverse("inventory:edit_vendor_price", args=[vendor_price.pk]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'value="50,25"')

        response = self.client.get(reverse("inventory:manage_vendor_prices"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "50,25")
        self.assertContains(response, "50.25")


class InventoryViewBoundaryTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="inventory-view-user",
            password="testpass123",
        )
        cls.vendor = Vendor.objects.create(
            name="Boundary Vendor",
            email="boundary@example.com",
            whatsapp_number="1111111111",
            is_active=True,
        )
        cls.product = Product.objects.create(
            product_name="Flour",
            pack_type="bag",
            quantity_per_pack="10",
            quantity_unit="Kg",
            display_name="Flour 10 Kg",
            is_active=True,
        )
        cls.vendor_price = VendorProductPrice.objects.create(
            product=cls.product,
            vendor=cls.vendor,
            price=Decimal("12.50"),
            currency="EUR",
            is_active=True,
        )

    def setUp(self):
        self.client.force_login(self.user)

    def test_product_list_still_shows_cart_quantity(self):
        cart = Cart.objects.create(user=self.user, status=Cart.STATUS_OPEN)
        CartItem.objects.create(
            cart=cart,
            product=self.product,
            vendor=self.vendor,
            quantity=2,
            unit_price=self.vendor_price.price,
        )

        response = self.client.get(reverse("inventory:product_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "In cart: 2")

    def test_price_tracker_still_renders_under_inventory(self):
        response = self.client.get(reverse("inventory:price_tracker"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Price Tracker")
        self.assertContains(response, self.product.display_name)


class ProductBulkImportTests(TestCase):
    workbook_headers = [
        "Product Name",
        "Display Name",
        "Packing Type",
        "Quantity",
        "Unit",
        "Vendor Name",
        "Price",
        "Alternative Names",
    ]

    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="bulk-import-user",
            password="testpass123",
        )

    def setUp(self):
        self.client.force_login(self.user)

    def _upload_workbook(self, rows, *, filename="products.xlsx"):
        return SimpleUploadedFile(
            filename,
            _build_test_workbook(rows),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    def _upload_raw_file(self, content, *, filename, content_type="application/octet-stream"):
        return SimpleUploadedFile(
            filename,
            content,
            content_type=content_type,
        )

    def test_bulk_import_creates_new_records_and_aliases(self):
        response = self.client.post(
            reverse("inventory:import_products_workbook"),
            data={
                "workbook": self._upload_workbook(
                    [
                        self.workbook_headers,
                        [
                            "Basmati Rice",
                            "Basmati Rice 5 Kg Bag",
                            "Bag",
                            5,
                            "Kg",
                            "Spice Vendor",
                            12.5,
                            "Rice Premium, Rice Bag",
                        ],
                    ]
                ),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Import Summary")
        self.assertContains(response, "Products Created")
        self.assertContains(response, ">1<", html=False)

        product = Product.objects.get(product_name="Basmati Rice")
        vendor = Vendor.objects.get(name="Spice Vendor")
        vendor_price = VendorProductPrice.objects.get(product=product, vendor=vendor)

        self.assertEqual(product.display_name, "Basmati Rice 5 Kg Bag")
        self.assertEqual(product.pack_type, "bag")
        self.assertEqual(product.quantity_per_pack, Decimal("5.00"))
        self.assertEqual(vendor_price.price, Decimal("12.50"))
        self.assertEqual(
            list(product.aliases.order_by("alias_name").values_list("alias_name", flat=True)),
            ["Rice Bag", "Rice Premium"],
        )

    def test_valid_openpyxl_generated_xlsx_is_accepted(self):
        response = self.client.post(
            reverse("inventory:import_products_workbook"),
            data={
                "workbook": self._upload_workbook(
                    [
                        self.workbook_headers,
                        [
                            "Cinnamon",
                            "Cinnamon 500 G Pack",
                            "Pack",
                            0.5,
                            "Kg",
                            "Excel Vendor",
                            9.25,
                            "",
                        ],
                    ]
                ),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Import Summary")
        self.assertTrue(Product.objects.filter(product_name="Cinnamon").exists())

    def test_bulk_import_extracts_quantity_and_infers_unit_from_safe_quantity_text(self):
        response = self.client.post(
            reverse("inventory:import_products_workbook"),
            data={
                "workbook": self._upload_workbook(
                    [
                        self.workbook_headers,
                        ["Rice A", "Rice A 10 Kg Bag", "Bag", "10kg", "", "Vendor A", 10, ""],
                        ["Rice B", "Rice B 5 Kg Bag", "Bag", "5 kg", "", "Vendor B", 11, ""],
                        ["Oil A", "Oil A 1 L Tin", "Tin", "1ltr", "", "Vendor C", 12, ""],
                        ["Oil B", "Oil B 15 L Tin", "Tin", "15 l", "", "Vendor D", 13, ""],
                        ["Flour A", "Flour A 10 Kg Bag", "Bag", "10,0", "kg", "Vendor E", 14, ""],
                    ]
                ),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Import Summary")

        rice_a = Product.objects.get(product_name="Rice A")
        rice_b = Product.objects.get(product_name="Rice B")
        oil_a = Product.objects.get(product_name="Oil A")
        oil_b = Product.objects.get(product_name="Oil B")
        flour_a = Product.objects.get(product_name="Flour A")

        self.assertEqual(rice_a.quantity_per_pack, Decimal("10.00"))
        self.assertEqual(rice_a.quantity_unit, "Kg")
        self.assertEqual(rice_b.quantity_per_pack, Decimal("5.00"))
        self.assertEqual(rice_b.quantity_unit, "Kg")
        self.assertEqual(oil_a.quantity_per_pack, Decimal("1.00"))
        self.assertEqual(oil_a.quantity_unit, "L")
        self.assertEqual(oil_b.quantity_per_pack, Decimal("15.00"))
        self.assertEqual(oil_b.quantity_unit, "L")
        self.assertEqual(flour_a.quantity_per_pack, Decimal("10.00"))
        self.assertEqual(flour_a.quantity_unit, "Kg")

    def test_bulk_import_prefers_explicit_unit_column_when_quantity_also_contains_unit(self):
        response = self.client.post(
            reverse("inventory:import_products_workbook"),
            data={
                "workbook": self._upload_workbook(
                    [
                        self.workbook_headers,
                        ["Sugar", "Sugar 1 Kg Pack", "Pack", "1ltr", "Kg", "Vendor", 5, ""],
                    ]
                ),
            },
        )

        self.assertEqual(response.status_code, 200)
        product = Product.objects.get(product_name="Sugar")
        self.assertEqual(product.quantity_per_pack, Decimal("1.00"))
        self.assertEqual(product.quantity_unit, "Kg")

    def test_bulk_import_reuses_existing_product_with_normalized_duplicate_key(self):
        product = Product.objects.create(
            product_name="Olive Oil",
            pack_type="tin",
            quantity_per_pack="10",
            quantity_unit="L",
            display_name="Olive Oil 10 L Tin",
            is_active=True,
        )

        response = self.client.post(
            reverse("inventory:import_products_workbook"),
            data={
                "workbook": self._upload_workbook(
                    [
                        self.workbook_headers,
                        [
                            "  olive oil ",
                            "Do Not Overwrite",
                            "Tin",
                            "10.00",
                            " l ",
                            "Vendor One",
                            18,
                            "Oil Tin",
                        ],
                    ]
                ),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(Product.objects.count(), 1)
        product.refresh_from_db()
        self.assertEqual(product.pk, Product.objects.get().pk)
        self.assertEqual(product.display_name, "Olive Oil 10 L Tin")
        self.assertEqual(
            list(product.aliases.values_list("alias_name", flat=True)),
            ["Oil Tin"],
        )

    def test_bulk_import_reuses_existing_vendor_by_normalized_name(self):
        Vendor.objects.create(
            name="Fresh Foods",
            email="",
            whatsapp_number="",
            is_active=True,
        )

        response = self.client.post(
            reverse("inventory:import_products_workbook"),
            data={
                "workbook": self._upload_workbook(
                    [
                        self.workbook_headers,
                        [
                            "Cardamom",
                            "Cardamom 1 Kg Pack",
                            "Pack",
                            1,
                            "Kg",
                            " fresh foods ",
                            "25.00",
                            "",
                        ],
                    ]
                ),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(Vendor.objects.count(), 1)
        self.assertEqual(Vendor.objects.get().name, "Fresh Foods")

    def test_bulk_import_updates_existing_vendor_price(self):
        vendor = Vendor.objects.create(
            name="Bulk Vendor",
            email="",
            whatsapp_number="",
            is_active=True,
        )
        product = Product.objects.create(
            product_name="Garam Masala",
            pack_type="pack",
            quantity_per_pack="1",
            quantity_unit="Kg",
            display_name="Garam Masala 1 Kg Pack",
            is_active=True,
        )
        VendorProductPrice.objects.create(
            product=product,
            vendor=vendor,
            price=Decimal("15.00"),
            is_active=True,
        )

        response = self.client.post(
            reverse("inventory:import_products_workbook"),
            data={
                "workbook": self._upload_workbook(
                    [
                        self.workbook_headers,
                        [
                            "Garam Masala",
                            "Garam Masala 1 Kg Pack",
                            "Pack",
                            1,
                            "Kg",
                            "Bulk Vendor",
                            "18,75",
                            "",
                        ],
                    ]
                ),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Vendor Prices Updated")

        vendor_price = VendorProductPrice.objects.get(product=product, vendor=vendor)
        self.assertEqual(vendor_price.price, Decimal("18.75"))

    def test_bulk_import_rejects_missing_headers(self):
        response = self.client.post(
            reverse("inventory:import_products_workbook"),
            data={
                "workbook": self._upload_workbook(
                    [
                        [
                            "Product Name",
                            "Display Name",
                            "Packing Type",
                            "Quantity",
                            "Unit",
                            "Vendor Name",
                            "Price",
                        ],
                        ["Rice", "Rice 5 Kg Bag", "Bag", 5, "Kg", "Vendor", 10],
                    ]
                ),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Missing required headers: Alternative Names.")
        self.assertFalse(Product.objects.exists())
        self.assertFalse(Vendor.objects.exists())
        self.assertFalse(VendorProductPrice.objects.exists())

    def test_bulk_import_rejects_invalid_non_xlsx_file(self):
        response = self.client.post(
            reverse("inventory:import_products_workbook"),
            data={
                "workbook": self._upload_raw_file(
                    b"plain text,not an excel workbook",
                    filename="products.csv",
                    content_type="text/csv",
                ),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Upload an .xlsx Excel file.")
        self.assertFalse(Product.objects.exists())

    def test_bulk_import_shows_clean_error_for_corrupt_xlsx_file(self):
        response = self.client.post(
            reverse("inventory:import_products_workbook"),
            data={
                "workbook": self._upload_raw_file(
                    b"not really an xlsx workbook",
                    filename="broken.xlsx",
                ),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Upload a valid .xlsx workbook.")
        self.assertFalse(Product.objects.exists())

    def test_bulk_import_still_rejects_ambiguous_quantity_values(self):
        response = self.client.post(
            reverse("inventory:import_products_workbook"),
            data={
                "workbook": self._upload_workbook(
                    [
                        self.workbook_headers,
                        ["Chilli", "Chilli Pack", "Pack", "about 5 kg", "", "Vendor", 5, ""],
                    ]
                ),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Quantity must be a valid number greater than zero.")
        self.assertContains(response, "Unit is required.")
        self.assertFalse(Product.objects.exists())

    def test_bulk_import_skips_invalid_rows_and_reports_reasons(self):
        response = self.client.post(
            reverse("inventory:import_products_workbook"),
            data={
                "workbook": self._upload_workbook(
                    [
                        self.workbook_headers,
                        [
                            "Turmeric",
                            "Turmeric 1 Kg Pack",
                            "Pack",
                            1,
                            "Kg",
                            "Spice House",
                            7.5,
                            "",
                        ],
                        [
                            "Chilli Powder",
                            "",
                            "Pack",
                            1,
                            "Kg",
                            "",
                            5,
                            "",
                        ],
                    ]
                ),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(Product.objects.count(), 1)
        self.assertEqual(Vendor.objects.count(), 1)
        self.assertContains(response, "Rows Skipped")
        self.assertContains(response, "Row 3:")
        self.assertContains(response, "Display Name is required for a new product.")
        self.assertContains(response, "Vendor Name is required.")
