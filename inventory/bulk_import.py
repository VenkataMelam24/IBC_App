from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import re
import zipfile

from django.db import transaction
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .currency import DECIMAL_PRECISION, DEFAULT_CURRENCY
from .forms import ProductForm
from .models import Product, Vendor, VendorProductPrice
from .product_aliases import merge_product_aliases


REQUIRED_IMPORT_HEADERS = [
    "Product Name",
    "Display Name",
    "Packing Type",
    "Quantity",
    "Unit",
    "Vendor Name",
    "Price",
    "Alternative Names",
]

class BulkImportError(Exception):
    pass


_QUANTITY_WITH_UNIT_RE = re.compile(
    r"^(?P<number>\d+(?:[.,]\d+)?)\s*(?P<unit>[A-Za-z]+)\s*$"
)

_UNIT_NORMALIZATION_MAP = {
    "kg": "Kg",
    "kgs": "Kg",
    "kilo": "Kg",
    "kilos": "Kg",
    "kilogram": "Kg",
    "kilograms": "Kg",
    "g": "G",
    "gm": "G",
    "gms": "G",
    "gram": "G",
    "grams": "G",
    "l": "L",
    "lt": "L",
    "ltr": "L",
    "ltrs": "L",
    "litre": "L",
    "litres": "L",
    "liter": "L",
    "liters": "L",
    "ml": "Ml",
    "mls": "Ml",
    "millilitre": "Ml",
    "millilitres": "Ml",
    "milliliter": "Ml",
    "milliliters": "Ml",
    "pc": "Pcs",
    "pcs": "Pcs",
    "piece": "Pcs",
    "pieces": "Pcs",
}


def import_products_from_workbook(uploaded_file):
    workbook_rows = _read_workbook_rows(uploaded_file)
    summary = {
        "products_created": 0,
        "vendors_created": 0,
        "vendor_prices_added": 0,
        "vendor_prices_updated": 0,
        "rows_skipped": 0,
        "skipped_rows": [],
    }

    product_cache = _build_product_cache()
    vendor_cache = _build_vendor_cache()
    vendor_price_cache = {
        (vendor_price.vendor_id, vendor_price.product_id): vendor_price
        for vendor_price in VendorProductPrice.objects.select_related("vendor", "product")
    }
    display_name_cache = {
        _normalize_lookup(product.display_name): product
        for product in Product.objects.exclude(display_name="")
    }
    default_pack_type_map = ProductForm._default_pack_type_map()
    custom_pack_type_lookup = {
        pack_type.casefold(): pack_type
        for pack_type in ProductForm._existing_custom_pack_types()
    }

    for row_number, row in workbook_rows:
        try:
            result = _import_workbook_row(
                row=row,
                product_cache=product_cache,
                vendor_cache=vendor_cache,
                vendor_price_cache=vendor_price_cache,
                display_name_cache=display_name_cache,
                default_pack_type_map=default_pack_type_map,
                custom_pack_type_lookup=custom_pack_type_lookup,
            )
        except BulkImportError as exc:
            summary["rows_skipped"] += 1
            summary["skipped_rows"].append(
                {
                    "row_number": row_number,
                    "reason": str(exc),
                }
            )
            continue

        if result["product_created"]:
            summary["products_created"] += 1

        if result["vendor_created"]:
            summary["vendors_created"] += 1

        if result["vendor_price_created"]:
            summary["vendor_prices_added"] += 1

        if result["vendor_price_updated"]:
            summary["vendor_prices_updated"] += 1

    return summary


def _import_workbook_row(
    *,
    row,
    product_cache,
    vendor_cache,
    vendor_price_cache,
    display_name_cache,
    default_pack_type_map,
    custom_pack_type_lookup,
):
    product_name = _clean_text(row.get("Product Name"))
    display_name = _clean_text(row.get("Display Name"))
    pack_type = _clean_text(row.get("Packing Type"))
    quantity_raw = row.get("Quantity")
    quantity_unit = _clean_text(row.get("Unit"))
    vendor_name = _clean_text(row.get("Vendor Name"))
    price_raw = row.get("Price")
    alternative_names = _split_alternative_names(row.get("Alternative Names", ""))

    quantity, quantity_unit = _parse_quantity_and_unit(
        quantity_raw,
        quantity_unit,
    )

    errors = []
    if not product_name:
        errors.append("Product Name is required.")
    if not pack_type:
        errors.append("Packing Type is required.")
    if not quantity_unit:
        errors.append("Unit is required.")
    if not vendor_name:
        errors.append("Vendor Name is required.")

    if quantity is None:
        errors.append("Quantity must be a valid number greater than zero.")

    price = _parse_decimal(price_raw, "Price", allow_zero=True)
    if price is None:
        errors.append("Price must be a valid number.")

    canonical_pack_type = _canonicalize_pack_type(
        pack_type,
        default_pack_type_map=default_pack_type_map,
        custom_pack_type_lookup=custom_pack_type_lookup,
    )
    if not canonical_pack_type:
        errors.append("Packing Type is required.")

    product = None
    conflicting_product = None
    display_name_key = _normalize_lookup(display_name)
    if product_name and canonical_pack_type and quantity is not None and quantity_unit:
        product_key = _build_product_key(
            product_name=product_name,
            pack_type=canonical_pack_type,
            quantity=quantity,
            quantity_unit=quantity_unit,
        )
        product = product_cache.get(product_key)
        conflicting_product = display_name_cache.get(display_name_key) if display_name_key else None

        if product is None and not display_name:
            errors.append("Display Name is required for a new product.")

        if product is None and conflicting_product is not None:
            errors.append(
                f'Display Name "{display_name}" already belongs to another product.'
            )

    if errors:
        raise BulkImportError(" ".join(errors))

    product_key = _build_product_key(
        product_name=product_name,
        pack_type=canonical_pack_type,
        quantity=quantity,
        quantity_unit=quantity_unit,
    )

    product_created = False
    vendor_created = False
    vendor_price_created = False
    vendor_price_updated = False

    with transaction.atomic():
        if product is None:
            product = Product.objects.create(
                product_name=product_name,
                pack_type=canonical_pack_type,
                quantity_per_pack=quantity,
                quantity_unit=quantity_unit,
                display_name=display_name,
                is_active=True,
            )
            product_cache[product_key] = product
            display_name_cache[display_name_key] = product
            product_created = True
            if canonical_pack_type.casefold() not in default_pack_type_map:
                custom_pack_type_lookup[canonical_pack_type.casefold()] = canonical_pack_type
        elif not product.display_name and display_name:
            if conflicting_product is not None and conflicting_product.pk != product.pk:
                raise BulkImportError(
                    f'Display Name "{display_name}" already belongs to another product.'
                )
            product.display_name = display_name
            product.save(update_fields=["display_name"])
            display_name_cache[display_name_key] = product

        if alternative_names:
            merge_product_aliases(product, alternative_names)

        vendor_key = _normalize_lookup(vendor_name)
        vendor = vendor_cache.get(vendor_key)
        if vendor is None:
            vendor = Vendor.objects.create(
                name=vendor_name,
                email="",
                whatsapp_number="",
                is_active=True,
            )
            vendor_cache[vendor_key] = vendor
            vendor_created = True

        vendor_price_key = (vendor.pk, product.pk)
        vendor_price = vendor_price_cache.get(vendor_price_key)
        if vendor_price is None:
            vendor_price = VendorProductPrice(
                vendor=vendor,
                product=product,
                price=price,
                currency=DEFAULT_CURRENCY,
                is_active=True,
            )
            vendor_price.save()
            vendor_price_cache[vendor_price_key] = vendor_price
            vendor_price_created = True
        else:
            should_save = False
            if vendor_price.price != price:
                vendor_price.price = price
                vendor_price_updated = True
                should_save = True
            if not vendor_price.is_active:
                vendor_price.is_active = True
                should_save = True

            if should_save:
                vendor_price.save(update_fields=["price", "is_active"])

    return {
        "product_created": product_created,
        "vendor_created": vendor_created,
        "vendor_price_created": vendor_price_created,
        "vendor_price_updated": vendor_price_updated,
    }


def _build_product_cache():
    cache = {}
    for product in Product.objects.all():
        cache[
            _build_product_key(
                product_name=product.product_name,
                pack_type=product.pack_type,
                quantity=product.quantity_per_pack,
                quantity_unit=product.quantity_unit,
            )
        ] = product
    return cache


def _build_vendor_cache():
    cache = {}
    for vendor in Vendor.objects.all().order_by("id"):
        cache.setdefault(_normalize_lookup(vendor.name), vendor)
    return cache


def _build_product_key(*, product_name, pack_type, quantity, quantity_unit):
    return (
        _normalize_lookup(product_name),
        _normalize_lookup(pack_type),
        _quantize_decimal(quantity),
        _normalize_lookup(quantity_unit),
    )


def _canonicalize_pack_type(value, *, default_pack_type_map, custom_pack_type_lookup):
    cleaned_value = _clean_text(value)
    if not cleaned_value:
        return ""

    normalized_value = cleaned_value.casefold()
    if normalized_value in default_pack_type_map:
        return default_pack_type_map[normalized_value]

    existing_custom_value = custom_pack_type_lookup.get(normalized_value)
    if existing_custom_value:
        return existing_custom_value

    custom_pack_type_lookup[normalized_value] = cleaned_value
    return cleaned_value


def _split_alternative_names(raw_value):
    return [
        alias.strip()
        for alias in str(raw_value or "").split(",")
        if alias.strip()
    ]


def _clean_text(value):
    return " ".join(str(value or "").split()).strip()


def _normalize_lookup(value):
    return _clean_text(value).casefold()


def _quantize_decimal(value):
    return Decimal(str(value)).quantize(DECIMAL_PRECISION, rounding=ROUND_HALF_UP)


def _parse_decimal(raw_value, label, *, allow_zero):
    value = _clean_text(raw_value)
    if not value:
        return None

    normalized_value = value.replace(",", ".")
    if normalized_value.count(".") > 1:
        return None

    try:
        decimal_value = Decimal(normalized_value)
    except (InvalidOperation, ValueError):
        return None

    if decimal_value < 0 or (not allow_zero and decimal_value <= 0):
        return None

    return decimal_value.quantize(DECIMAL_PRECISION, rounding=ROUND_HALF_UP)


def _normalize_quantity_unit(value, *, allow_unknown):
    cleaned_value = _clean_text(value).rstrip(".")
    if not cleaned_value:
        return ""

    normalized_value = cleaned_value.casefold()
    canonical_unit = _UNIT_NORMALIZATION_MAP.get(normalized_value)
    if canonical_unit:
        return canonical_unit

    if allow_unknown:
        return cleaned_value

    return ""


def _parse_quantity_and_unit(raw_quantity, raw_unit):
    explicit_unit = _normalize_quantity_unit(raw_unit, allow_unknown=True)
    quantity = _parse_decimal(raw_quantity, "Quantity", allow_zero=False)
    if quantity is not None:
        return quantity, explicit_unit

    quantity_text = _clean_text(raw_quantity)
    if not quantity_text:
        return None, explicit_unit

    match = _QUANTITY_WITH_UNIT_RE.fullmatch(quantity_text)
    if match is None:
        return None, explicit_unit

    quantity = _parse_decimal(match.group("number"), "Quantity", allow_zero=False)
    inferred_unit = _normalize_quantity_unit(match.group("unit"), allow_unknown=False)

    if quantity is None:
        return None, explicit_unit

    if explicit_unit:
        return quantity, explicit_unit

    return quantity, inferred_unit


def _normalize_header(value):
    return _clean_text(value).casefold()


def _read_workbook_rows(uploaded_file):
    workbook = None
    try:
        if hasattr(uploaded_file, "seek"):
            uploaded_file.seek(0)
        workbook = load_workbook(
            filename=uploaded_file,
            read_only=True,
            data_only=True,
        )
    except (InvalidFileException, OSError, ValueError, zipfile.BadZipFile) as exc:
        raise BulkImportError("Upload a valid .xlsx workbook.") from exc
    try:
        worksheet = workbook.worksheets[0] if workbook.worksheets else None
        if worksheet is None:
            raise BulkImportError("The workbook does not contain any worksheets.")

        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise BulkImportError("The workbook does not contain any rows.")

        header_values = rows[0]
        header_map = {}
        normalized_required_headers = {
            _normalize_header(header): header
            for header in REQUIRED_IMPORT_HEADERS
        }

        for index, value in enumerate(header_values):
            canonical_header = normalized_required_headers.get(_normalize_header(value))
            if canonical_header:
                header_map[index] = canonical_header

        missing_headers = [
            header
            for header in REQUIRED_IMPORT_HEADERS
            if header not in set(header_map.values())
        ]
        if missing_headers:
            missing_header_text = ", ".join(missing_headers)
            raise BulkImportError(f"Missing required headers: {missing_header_text}.")

        parsed_rows = []
        for row_number, row_values in enumerate(rows[1:], start=2):
            row_data = {
                header: _clean_text(
                    row_values[index] if index < len(row_values) else ""
                )
                for index, header in header_map.items()
            }
            if not any(row_data.values()):
                continue
            parsed_rows.append((row_number, row_data))

        return parsed_rows
    finally:
        if workbook is not None:
            workbook.close()
        if hasattr(uploaded_file, "seek"):
            uploaded_file.seek(0)
